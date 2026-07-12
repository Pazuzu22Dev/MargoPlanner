import csv
import io
import json

from services.input_service import InputPayload


def _rows_to_text(rows):
    return json.dumps({"rows": rows}, ensure_ascii=False, default=str)


def extract_content(payload: InputPayload):
    if payload.source_type in {"text", "forwarded_message"}:
        return str(payload.content)
    if payload.source_type == "csv":
        text = payload.content.decode("utf-8-sig")
        return _rows_to_text(list(csv.DictReader(io.StringIO(text))))
    if payload.source_type == "xlsx":
        from openpyxl import load_workbook

        workbook = load_workbook(io.BytesIO(payload.content), data_only=True)
        tables = []
        for sheet in workbook.worksheets:
            values = list(sheet.iter_rows(values_only=True))
            if not values:
                continue
            headers = [str(value or f"column_{index + 1}") for index, value in enumerate(values[0])]
            rows = [dict(zip(headers, row)) for row in values[1:] if any(value is not None for value in row)]
            tables.append({"sheet": sheet.title, "rows": rows})
        return json.dumps({"tables": tables}, ensure_ascii=False, default=str)
    if payload.source_type == "pdf":
        from pypdf import PdfReader

        reader = PdfReader(io.BytesIO(payload.content))
        return "\n\n".join(page.extract_text() or "" for page in reader.pages)
    if payload.source_type == "image":
        return {"image": payload.content, "mime_type": payload.mime_type, "caption": payload.caption}
    raise ValueError("Этот формат пока нельзя прочитать")
